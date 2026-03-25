import ttkbootstrap as tb
from ttkbootstrap.constants import *
from database import cursor
import matplotlib.pyplot as plt
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
plt.style.use("dark_background")
from utils.dashboard_utils import get_type_stats
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib
matplotlib.use("TkAgg")

class Dashboard(tb.Frame):
    def __init__(self, master, user=None):
        self.user = user
        super().__init__(master)
        self.pack(fill=BOTH, expand=True)

        self.create_ui()

    def create_ui(self):
        # ---------- HEADER ----------
        header = tb.Frame(self)
        header.pack(fill=X, padx=25, pady=(15, 5))

        tb.Label(
            header,
            text=f"👋 Welcome, {self.user['username'].capitalize()}",
            font=("Segoe UI", 20, "bold")
        ).pack(anchor="w")

        tb.Label(
            header,
            text="Here's what's happening today",
            font=("Segoe UI", 10),
            bootstyle="secondary"
        ).pack(anchor="w")

        # ---------- STATS ----------
        stats_frame = tb.Frame(self)
        stats_frame.pack(fill=X, padx=20, pady=15)

        users = cursor.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        products = cursor.execute("SELECT COUNT(*) FROM batteries").fetchone()[0]
        sales = cursor.execute("SELECT COUNT(*) FROM sales").fetchone()[0]
        revenue = cursor.execute("SELECT SUM(total) FROM sales").fetchone()[0] or 0

        def card(parent, title, value, color, icon, prefix="",key=None):
            frame = tb.Frame(parent, padding=20, bootstyle=color)
            frame.pack(side=LEFT, padx=12, fill=X, expand=True)

            
            top = tb.Frame(frame, bootstyle=color)
            top.pack(fill=X)

            tb.Label(
                top,
                text=icon,
                font=("Segoe UI", 22),
                bootstyle=f"{color}-inverse"
            ).pack(side=LEFT)

            tb.Label(
                top,
                text=title,
                font=("Segoe UI", 14, "bold"),
                bootstyle=f"{color}-inverse"
            ).pack(side=LEFT, padx=10)   

            
            
            value_label = tb.Label(
                frame,
                text="0",   # start from 0
                font=("Segoe UI", 24, "bold"),
                bootstyle=f"{color}-inverse",
                anchor="w"
            )
            value_label.pack(fill=X, pady=(10, 0))
            if key:
                setattr(self, f"{key}_label", value_label)
            self.animate_counter(value_label, value, prefix)
             
        card(stats_frame, "Users", users, "success", "👤",key="users")
        card(stats_frame, "Products", products, "info", "🔋",key="products")
        card(stats_frame, "Sales", sales, "warning", "🧾",key="sales")
        card(stats_frame, "Revenue", revenue, "danger", "💰", "₹",key="revenue")
    
        main_frame = tb.Frame(self)
        

        #left panel
        left_panel = tb.Frame(main_frame)

        # -------- CHART WRAPPER -------- #
        chart_wrapper = tb.Frame(left_panel)
        chart_wrapper.pack(fill=BOTH, expand=True)

        # LEFT CHART (Sales)
        left_chart = tb.Frame(chart_wrapper)
        left_chart.pack(side=LEFT, fill=BOTH, expand=True, padx=5)

        sales_frame = tb.LabelFrame(left_chart, text="📊 Sales by Type")
        sales_frame.pack(fill=BOTH, expand=True)

        self.chart_container = tb.Frame(sales_frame)
        self.chart_container.pack(fill=BOTH, expand=True, padx=10, pady=10)

        # RIGHT CHART (Profit)
        right_chart = tb.Frame(chart_wrapper)
        right_chart.pack(side=LEFT, fill=BOTH, expand=True, padx=5)

        profit_frame = tb.LabelFrame(right_chart, text="📈 Profit Overview")
        profit_frame.pack(fill=BOTH, expand=True)

        self.profit_chart_container = tb.Frame(profit_frame)
        self.profit_chart_container.pack(fill=BOTH, expand=True, padx=10, pady=10)

        # -------- BATTERY TYPE STATS -------- #
        type_frame = tb.LabelFrame(left_panel, text="🔋 Battery Types")
        type_frame.pack(fill=X, padx=20, pady=10)

        inner = tb.Frame(type_frame, padding=10)
        inner.pack(fill=X)

        self.type_cards_frame = inner

        self.type_cards_frame = tb.Frame(type_frame)
        self.type_cards_frame.pack(fill=X)

        left_panel.pack(side=LEFT, fill=BOTH, expand=True)
        # your charts

        self.create_low_stock_panel(main_frame)
        main_frame.pack(fill=BOTH, expand=True, padx=20, pady=10)

        self.current_filter = "all"

        filter_frame = tb.Frame(self)
        filter_frame.pack(anchor="e", padx=20)

        def set_filter(val):
            self.current_filter = val

        tb.Button(filter_frame, text="All", command=lambda: setattr(self, "current_filter", "all")).pack(side=LEFT)
        tb.Button(filter_frame, text="Today", command=lambda: setattr(self, "current_filter", "today")).pack(side=LEFT)
        tb.Button(filter_frame, text="Week", command=lambda: setattr(self, "current_filter", "week")).pack(side=LEFT)
        tb.Button(filter_frame, text="Month", command=lambda: setattr(self, "current_filter", "month")).pack(side=LEFT)

        btn_frame = tb.Frame(self)
        btn_frame.pack(anchor="e", padx=20, pady=5)

        tb.Button(
            btn_frame,
            text="📦 Export Products",
            bootstyle="info",
            command=self.export_products_excel
        ).pack(side=LEFT, padx=5)

        tb.Button(
            btn_frame,
            text="💰 Export Sales",
            bootstyle="success",
            command=self.export_sales_excel
        ).pack(side=LEFT, padx=5)
        tb.Button(
            btn_frame,
            text="📊 Export Full Report",
            bootstyle="success",
            command=self.export_all_excel
        ).pack(side=LEFT, padx=5)

        
        self.load_type_stats()
        self.load_sales_chart()
        self.load_profit_chart()
        self.refresh_dashboard()

    def create_low_stock_panel(self, parent):
        right_panel = tb.Frame(parent, width=250, bootstyle="dark")
        right_panel.pack(side=RIGHT, fill=Y, padx=(10, 0))
        right_panel.pack_propagate(False)
        tb.Label(
            right_panel,
            text="⚠ Low Stock",
            font=("Segoe UI", 12, "bold"),
            bootstyle="danger"
        ).pack(anchor="w", padx=10, pady=10)
        import tkinter as tk
        canvas = tk.Canvas(right_panel, bg="#212529", highlightthickness=0)
        scrollbar = tb.Scrollbar(right_panel, orient="vertical", command=canvas.yview)
        self.low_stock_frame = tb.Frame(canvas)
        self.low_stock_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.low_stock_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        self.load_low_stock()
        self.auto_refresh_low_stock()
    def animate_counter(self, label, target, prefix="", duration=1000):
        try:
            target = float(target)
        except:
            label.config(text=f"{prefix}{target}")
            return

        steps = 40
        delay = duration // steps

        def ease_out(t):
            return 1 - (1 - t) ** 3   # cubic easing

        def format_value(val):
            val = int(val)
            s = str(val)
            if len(s) <= 3:
                return f"{prefix}{s}"
            else:
                last3 = s[-3:]
                rest = s[:-3]
                rest = ",".join([rest[max(i-2,0):i] for i in range(len(rest), 0, -2)][::-1])
                return f"{prefix}{rest},{last3}" 

        def update(step=0):
            if step > steps:
                label.config(text=format_value(target))
                return

            progress = step / steps
            eased = ease_out(progress)
            current = target * eased

            label.config(text=format_value(current))
            label.after(delay, update, step + 1)

        update()
    def refresh_dashboard(self):
        # Re-fetch data
        users = cursor.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        products = cursor.execute("SELECT COUNT(*) FROM batteries").fetchone()[0]
        sales = cursor.execute("SELECT COUNT(*) FROM sales").fetchone()[0]
        revenue = cursor.execute("SELECT SUM(total) FROM sales").fetchone()[0] or 0

        # Re-run animations
        self.animate_counter(self.users_label, users)
        self.animate_counter(self.products_label, products)
        self.animate_counter(self.sales_label, sales)
        self.animate_counter(self.revenue_label, revenue, "₹")
        self.load_type_stats()
        if hasattr(self, "sales_ax"):
            self.update_sales_chart()

        if hasattr(self, "profit_ax"):
            self.update_profit_chart()
        # Refresh every 10 sec
        self.after(10000, self.refresh_dashboard)
    def load_type_stats(self):
        # Clear old UI
        for widget in self.type_cards_frame.winfo_children():
            widget.destroy()

        stats = get_type_stats()

        if not stats:
            tb.Label(self.type_cards_frame, text="No data").pack()
            return

        for row in stats:
            box = tb.Frame(self.type_cards_frame, padding=10, bootstyle="secondary")
            box.pack(side=LEFT, padx=10, pady=5, expand=True, fill=X)

            tb.Label(
                box,
                text=row["type"],
                font=("Segoe UI", 12, "bold")
            ).pack(anchor="w")

            tb.Label(
                box,
                text=f"Items: {row['total_items']}",
                font=("Segoe UI", 10)
            ).pack(anchor="w")

            tb.Label(
                box,
                text=f"Stock: {row['total_qty']}",
                font=("Segoe UI", 10)
            ).pack(anchor="w")
    
    def get_sales_query(self):
        if self.current_filter == "today":
            return "SELECT * FROM sales WHERE DATE(date)=DATE('now')"

        elif self.current_filter == "week":
            return "SELECT * FROM sales WHERE DATE(date)>=DATE('now','-7 days')"

        elif self.current_filter == "month":
            return "SELECT * FROM sales WHERE strftime('%Y-%m',date)=strftime('%Y-%m','now')"

        else:
            return "SELECT * FROM sales"
    def get_sales_by_type(self):
        query = """
        SELECT b.type, SUM(si.quantity) as total
        FROM sale_items si
        JOIN batteries b ON si.product_id = b.id
        GROUP BY b.type
        """
        cursor.execute(query)
        return cursor.fetchall()
    def load_sales_chart(self):
        if hasattr(self, "sales_ax"):
            self.update_sales_chart()
            return

        fig, ax = plt.subplots(figsize=(3, 2), dpi=100)

        fig.patch.set_facecolor("#212529")
        ax.set_facecolor("#212529")

        self.sales_fig = fig
        self.sales_ax = ax

        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        self.sales_canvas = FigureCanvasTkAgg(fig, master=self.chart_container)
        self.sales_canvas.get_tk_widget().pack(fill="both", expand=True)

        self.update_sales_chart()
    def update_sales_chart(self):
        data = self.get_sales_by_type()

        if not data:
            return

        types = [row["type"] for row in data]
        values = [float(row["total"] or 0) for row in data]

        ax = self.sales_ax
        ax.clear()

        ax.set_facecolor("#212529")
        ax.bar(types, values, color="#0dcaf0", edgecolor="#0dcaf0")

        ax.set_title("Sales by Type", color="white", fontsize=10)
        ax.tick_params(colors="white", labelsize=8)

        for spine in ax.spines.values():
            spine.set_color("#444")

        self.sales_canvas.draw_idle()
    
    def get_profit_chart_data(self):
        query = """
        SELECT DATE(s.sold_date) as date,
            SUM(s.total) as revenue,
            SUM(si.quantity * b.purchase_price) as cost
        FROM sales s
        JOIN sale_items si ON s.id = si.sale_id
        JOIN batteries b ON si.product_id = b.id
        GROUP BY DATE(s.sold_date)
        ORDER BY DATE(s.sold_date)
        """
        cursor.execute(query)
        return cursor.fetchall()
    def load_profit_chart(self):
        if hasattr(self, "profit_ax"):
            self.update_profit_chart()
            return

        fig, ax = plt.subplots(figsize=(3, 2), dpi=100)

        fig.patch.set_facecolor("#212529")
        ax.set_facecolor("#212529")

        self.profit_fig = fig
        self.profit_ax = ax

        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        self.profit_canvas = FigureCanvasTkAgg(fig, master=self.profit_chart_container)
        self.profit_canvas.get_tk_widget().pack(fill="both", expand=True)

        self.update_profit_chart()
    def update_profit_chart(self):
        data = self.get_profit_chart_data()

        if not data:
            return

        dates = [row["date"] for row in data]
        revenue = [float(row["revenue"] or 0) for row in data]
        cost = [float(row["cost"] or 0) for row in data]
        profit = [r - c for r, c in zip(revenue, cost)]

        ax = self.profit_ax
        ax.clear()  

        ax.set_facecolor("#212529")

        ax.plot(dates, revenue, marker="o", linewidth=2, color="#28a745", label="Revenue")
        ax.plot(dates, profit, marker="o", linewidth=2, color="#dc3545", label="Profit")

        ax.set_title("Revenue vs Profit", color="white", fontsize=10)
        ax.tick_params(colors="white", labelsize=8)
        ax.legend(facecolor="#212529", edgecolor="#444", labelcolor="white")

        for spine in ax.spines.values():
            spine.set_color("#444")

        self.profit_canvas.draw_idle()  
    
    def export_sales_excel(self):
        file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )

        if not file:
            return

        query = self.get_sales_query() 
        data = cursor.execute(query).fetchall()
        headers = [col[0] for col in cursor.description]

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in data:
            ws.append(list(row))

        # auto width
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        wb.save(file)

        messagebox.showinfo(
            "Export",
            f"Sales exported ({self.current_filter}) ✅"
        )
    def export_products_excel(self):
        file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Products Excel"
        )

        if not file:
            return

        data = cursor.execute("SELECT * FROM batteries").fetchall()
        headers = [col[0] for col in cursor.description]

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Header
        ws.append(headers)

        # Data
        for row in data:
            ws.append(list(row))

        wb.save(file)
        messagebox.showinfo("Success", "Products exported to Excel ✅")
    def export_all_excel(self):
        file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Report"
        )

        if not file:
            return

        wb = Workbook()

        # ---------- PRODUCTS SHEET ----------
        ws1 = wb.active
        ws1.title = "Products"

        data1 = cursor.execute("SELECT * FROM batteries").fetchall()
        headers1 = [col[0] for col in cursor.description]

        ws1.append(headers1)
        for cell in ws1[1]:
            cell.font = Font(bold=True)

        for row in data1:
            ws1.append(list(row))

        # ---------- SALES SHEET ----------
        ws2 = wb.create_sheet(title="Sales")

        data2 = cursor.execute("SELECT * FROM sales").fetchall()
        headers2 = [col[0] for col in cursor.description]

        ws2.append(headers2)
        for cell in ws2[1]:
            cell.font = Font(bold=True)

        for row in data2:
            ws2.append(list(row))

        # ---------- AUTO WIDTH ----------
        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

        wb.save(file)

        messagebox.showinfo("Export", "Full report exported ✅")
    
    def load_low_stock(self):
        # Clear old items
        for widget in self.low_stock_frame.winfo_children():
            widget.destroy()

        low_stock = cursor.execute("""
            SELECT name, quantity
            FROM batteries
            WHERE quantity <= 5
        """).fetchall()

        if low_stock:
            for name, qty in low_stock:
                max_qty = max(qty, 10)   # adjust based on your system

                if qty <= 2:
                    style = "danger"
                elif qty <= 5:
                    style = "warning"
                else:
                    style = "success"

                item_frame = tb.Frame(self.low_stock_frame)
                item_frame.pack(fill=X, padx=8, pady=6)

                # Product name
                tb.Label(
                    item_frame,
                    text=f"🔋 {name}",
                    font=("Segoe UI", 9, "bold")
                ).pack(anchor="w")

                # Progress bar
                progress = tb.Progressbar(
                    item_frame,
                    maximum=max_qty,
                    value=qty,
                    bootstyle=style
                )
                progress.pack(fill=X, pady=2)

                # Quantity text
                tb.Label(
                    item_frame,
                    text=f"{qty} remaining",
                    font=("Segoe UI", 8)
                ).pack(anchor="w")
        else:
            tb.Label(
                self.low_stock_frame,
                text="No low stock items",
                font=("Segoe UI", 9)
            ).pack(pady=10)
    
    def auto_refresh_low_stock(self):
        self.load_low_stock()
        self.after(10000, self.auto_refresh_low_stock)